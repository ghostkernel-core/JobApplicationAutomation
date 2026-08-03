"""Build-time duplicate detection: have I already applied to this role?

This is the layer that reaches outside the watcher's own database and into the
workspace itself. It matters because the folder tree and the tracker workbook
contain years of applications made before this system existed — a fresh
database knows nothing about them, and without this check the first poll would
happily rebuild an application submitted last month.

Two independent sources are consulted and either one is sufficient:

    the folder tree      <YYYY>/<Company>/<YYYY-MM-DD> - <Role>/
    the tracker workbook <YYYY>/Job Applications Tracker <YYYY>.xlsx

The folder tree is authoritative for "work was done"; the tracker is
authoritative for "it was logged". They usually agree. When they disagree the
folder wins, because a missing tracker row is a known non-fatal outcome of the
pipeline (CLAUDE.md step 10).
"""

from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass
from pathlib import Path

from .config import REPO_ROOT, load_identity
from .normalize import company_key, title_similarity

# "2026-07-24 - AI Engineer"
FOLDER_PATTERN = re.compile(r"^(\d{4}-\d{2}-\d{2})\s*-\s*(.+)$")
YEAR_DIR = re.compile(r"^(19|20)\d{2}$")


@dataclass(frozen=True)
class ExistingApplication:
    company: str
    title: str
    applied_on: dt.date | None
    folder: str
    similarity: float
    origin: str  # "folder" or "tracker"

    def describe(self) -> str:
        when = self.applied_on.isoformat() if self.applied_on else "unknown date"
        where = self.folder or "(no folder recorded)"
        return f"{self.company} — {self.title} · applied {when} · {where}"


def _year_dirs(root: Path) -> list[Path]:
    return sorted(
        (p for p in root.iterdir() if p.is_dir() and YEAR_DIR.match(p.name)),
        reverse=True,
    )


def _scan_folders(company: str, title: str, cutoff: dt.date, ratio: float,
                  root: Path) -> list[ExistingApplication]:
    wanted = company_key(company)
    hits: list[ExistingApplication] = []
    for year_dir in _year_dirs(root):
        for company_dir in year_dir.iterdir():
            if not company_dir.is_dir() or company_key(company_dir.name) != wanted:
                continue
            for role_dir in company_dir.iterdir():
                if not role_dir.is_dir():
                    continue
                match = FOLDER_PATTERN.match(role_dir.name)
                if not match:
                    continue
                try:
                    applied = dt.date.fromisoformat(match.group(1))
                except ValueError:
                    continue
                if applied < cutoff:
                    continue
                similarity = title_similarity(title, match.group(2))
                if similarity >= ratio:
                    hits.append(ExistingApplication(
                        company=company_dir.name, title=match.group(2),
                        applied_on=applied, folder=str(role_dir),
                        similarity=similarity, origin="folder",
                    ))
    return hits


def _scan_tracker(company: str, title: str, cutoff: dt.date, ratio: float,
                  root: Path) -> list[ExistingApplication]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    wanted = company_key(company)
    hits: list[ExistingApplication] = []
    for year_dir in _year_dirs(root):
        path = year_dir / f"Job Applications Tracker {year_dir.name}.xlsx"
        if not path.exists():
            continue
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook["Applications"]
        except Exception:
            continue  # a locked or malformed workbook must not block a build
        try:
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            col = {str(v).strip(): i for i, v in enumerate(header_row) if v}
            need = ("Company", "Position Applied", "Date Applied")
            if not all(k in col for k in need):
                continue
            for row in sheet.iter_rows(min_row=2, values_only=True):
                raw_company = row[col["Company"]] if col["Company"] < len(row) else None
                if not raw_company or company_key(str(raw_company)) != wanted:
                    continue
                position = str(row[col["Position Applied"]] or "")
                applied = row[col["Date Applied"]]
                if isinstance(applied, dt.datetime):
                    applied = applied.date()
                if isinstance(applied, dt.date) and applied < cutoff:
                    continue
                similarity = title_similarity(title, position)
                if similarity >= ratio:
                    folder_idx = col.get("Application Folder")
                    folder = ""
                    if folder_idx is not None and folder_idx < len(row):
                        folder = str(row[folder_idx] or "")
                    hits.append(ExistingApplication(
                        company=str(raw_company), title=position,
                        applied_on=applied if isinstance(applied, dt.date) else None,
                        folder=folder, similarity=similarity, origin="tracker",
                    ))
        finally:
            workbook.close()
    return hits


def collect_existing(company: str, title: str, lookback_days: int = 365,
                     ratio: float = 0.8, root: Path | None = None
                     ) -> list[ExistingApplication]:
    """Every prior application matching this company+role, best first."""
    root = root or REPO_ROOT
    cutoff = dt.date.today() - dt.timedelta(days=lookback_days)
    hits = _scan_folders(company, title, cutoff, ratio, root)
    hits += _scan_tracker(company, title, cutoff, ratio, root)
    return sorted(hits, key=lambda h: (h.similarity, h.origin == "folder",
                                       h.applied_on or dt.date.min), reverse=True)


def find_existing(company: str, title: str, lookback_days: int = 365,
                  ratio: float = 0.8, root: Path | None = None
                  ) -> ExistingApplication | None:
    """Closest prior application to this company+role, or None.

    Folder hits are preferred over tracker hits at equal similarity, since a
    folder proves the documents exist.
    """
    hits = collect_existing(company, title, lookback_days, ratio, root)
    return hits[0] if hits else None


# --------------------------------------------------------------------------
# completeness
# --------------------------------------------------------------------------

# The two documents every run produces, English-only or not. German and
# Interview Prep are conditional, so they cannot be part of the test.
REQUIRED_PDFS = tuple(
    load_identity().doc_name(label, ".pdf") for label in ("CV", "Cover Letter")
)

OPTIONAL_PDFS = tuple(
    load_identity().doc_name(label, ".pdf")
    for label in ("Lebenslauf", "Anschreiben", "Interview Prep")
)


def missing_documents(folder: str | Path) -> list[str]:
    """Which required PDFs are absent from a deliverable folder.

    A folder alone is not proof that an application exists.
    `2026/Synergeticon/2026-07-13 - AI Engineer` holds no PDFs at all — an
    abandoned run. Treating that as "already applied" would block the role from
    ever being retried, so the duplicate check asks for the documents, not the
    directory.
    """
    path = Path(folder)
    if not path.is_dir():
        return list(REQUIRED_PDFS)
    present = {entry.name for entry in path.iterdir() if entry.is_file()}
    return [name for name in REQUIRED_PDFS if name not in present]


def present_documents(folder: str | Path) -> list[str]:
    """Deliverables actually in the folder, for the build-result message."""
    path = Path(folder)
    if not path.is_dir():
        return []
    present = {entry.name for entry in path.iterdir() if entry.is_file()}
    return [name for name in REQUIRED_PDFS + OPTIONAL_PDFS if name in present]


def is_complete(existing: ExistingApplication) -> bool:
    """Whether a prior application is finished enough to block a rebuild.

    Tracker rows without a folder are taken at their word: the tracker is only
    written at step 10, after final QA passed, so a row is stronger evidence
    than a directory.
    """
    if not existing.folder:
        return existing.origin == "tracker"
    return not missing_documents(existing.folder)
