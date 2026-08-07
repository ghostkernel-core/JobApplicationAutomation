r"""Rewrite already-stored paths into the workspace-relative form.

`scripts/workspace_paths.py` explains the convention and every writer now
follows it. This is the one-off pass that repairs what was written before them.

It exists because the workspace moved. Everything persisted at `D:\Job
Applications` still names that drive: 39 of 43 tracker rows, all 20
`builds.folder` values, all 22 `builds.log_path` values. A tracker row whose
folder does not resolve looks to `dedupe.is_complete` like an application with
no PDFs, so a finished application stops blocking a rebuild; a build log that
does not resolve is a progress message that can never be replayed.

Two stores, one convention, so one pass over both:

    <YYYY>/Job Applications Tracker <YYYY>.xlsx   column "Application Folder"
    automation/state/watch.db                     builds.folder, builds.log_path,
                                                  questions.folder

Only values that actually change are written, and a value `to_relative` cannot
read is left exactly as it is — visibly wrong beats silently mangled. Running it
twice is a no-op, so it is safe to re-run after any restore or merge.

Not touched: `builds.detail`, which holds a handful of messages with a path
embedded in prose. That is text written at the time, not a path field.

Usage:
    python scripts/normalize_stored_paths.py --dry-run   # report, write nothing
    python scripts/normalize_stored_paths.py             # workbooks + watcher db
    python scripts/normalize_stored_paths.py --year 2026 # that workbook alone
"""

from __future__ import annotations

import argparse
import sqlite3
import sys
from dataclasses import dataclass, field
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from workspace_paths import ROOT, YEAR_RE, looks_absolute, to_relative  # noqa: E402

FOLDER_HEADER = "Application Folder"

# (table, column) for every path the watcher database stores. Hardcoded, so the
# names interpolated into the UPDATE below are literals, never input.
DB_COLUMNS = (("builds", "folder"), ("builds", "log_path"), ("questions", "folder"))


@dataclass
class Result:
    """What one store turned out to hold."""
    label: str
    changed: list[str] = field(default_factory=list)
    left_alone: list[str] = field(default_factory=list)
    already_relative: int = 0

    @property
    def touched(self) -> bool:
        return bool(self.changed)


def _classify(where: str, old: str, root: Path) -> tuple[str, str | None, str | None]:
    r"""(new_value, changed_line, left_alone_line) for one stored value.

    A value is "left alone" when `to_relative` returned it unchanged *and* it is
    still absolute — rule 4 fell through, and we have no idea what it means. An
    unchanged relative value is simply already correct. "Absolute" here means
    absolute anywhere, not on the machine running the repair: a workbook copied
    onto Linux still holds `D:\...` rows, and those are the ones to leave alone.

    Either outcome may name something that is not on disk. That is worth saying
    and is never worth refusing over: a folder cleaned up months ago should
    still be recorded the portable way.
    """
    new = to_relative(old, root)
    missing = "" if (root / new).exists() else "  (names nothing on disk)"
    if new == old:
        if looks_absolute(old):
            return old, None, f"{where}: {old} -- could not be read; left alone"
        return old, None, None
    return new, f"{where}: {old} -> {new}{missing}", None


# --------------------------------------------------------------------------
# the tracker workbooks
# --------------------------------------------------------------------------

def find_workbooks(root: Path, year: int | None = None) -> list[Path]:
    if year is not None:
        path = root / str(year) / f"Job Applications Tracker {year}.xlsx"
        return [path] if path.exists() else []
    found = []
    for year_dir in sorted(p for p in root.iterdir()
                           if p.is_dir() and YEAR_RE.match(p.name)):
        path = year_dir / f"Job Applications Tracker {year_dir.name}.xlsx"
        if path.exists():
            found.append(path)
    return found


def normalize_workbook(path: Path, root: Path, dry_run: bool) -> Result:
    """Rewrite the Application Folder column of one tracker workbook.

    One cell's `.value` at a time. Unlike `cleanup_application.remove_tracker_row`
    nothing here shifts rows, so the per-row `Days Since Applied` formulas, the
    validation, and the conditional formatting are all left exactly as they were.
    """
    result = Result(label=str(path.relative_to(root)))
    try:
        from openpyxl import load_workbook
    except ImportError:
        result.left_alone.append("openpyxl is not installed; workbook skipped")
        return result

    workbook = load_workbook(path)
    try:
        if "Applications" not in workbook.sheetnames:
            result.left_alone.append("no Applications sheet; skipped")
            return result
        sheet = workbook["Applications"]
        column = next((c.column for c in sheet[1]
                       if str(c.value or "").strip() == FOLDER_HEADER), None)
        if column is None:
            result.left_alone.append(f"no {FOLDER_HEADER!r} column; skipped")
            return result

        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=column)
            old = str(cell.value or "").strip()
            if not old:
                continue
            new, changed, left = _classify(f"row {row}", old, root)
            if changed:
                result.changed.append(changed)
                if not dry_run:
                    cell.value = new
            elif left:
                result.left_alone.append(left)
            else:
                result.already_relative += 1

        if result.changed and not dry_run:
            workbook.save(path)
    finally:
        workbook.close()
    return result


# --------------------------------------------------------------------------
# the watcher database
# --------------------------------------------------------------------------

def running_watchers() -> list[int]:
    """Pids of live watcher instances, or an empty list if none / unknowable.

    `watcherctl` finds them by command line rather than by a pid file, and
    imports nothing from `watcher`, so this costs one process listing and works
    under any interpreter. If it cannot be imported at all, say nothing is
    running: a migration blocked by a broken import helps no one.
    """
    sys.path.insert(0, str(ROOT / "automation"))
    try:
        from watcherctl import find_watchers
        return [proc.pid for proc in find_watchers()]
    except Exception:
        return []


def normalize_database(path: Path, root: Path, dry_run: bool) -> Result:
    """Rewrite all three path columns of the watcher database."""
    result = Result(label=str(path.relative_to(root)))
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    try:
        tables = {r["name"] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type = 'table'")}
        for table, column in DB_COLUMNS:
            if table not in tables:
                result.left_alone.append(f"no {table} table; skipped")
                continue
            rows = conn.execute(
                f"SELECT id, {column} AS value FROM {table} "
                f"WHERE {column} IS NOT NULL AND {column} != ''").fetchall()
            for row in rows:
                old = str(row["value"]).strip()
                if not old:
                    continue
                where = f"{table}.{column} id {row['id']}"
                new, changed, left = _classify(where, old, root)
                if changed:
                    result.changed.append(changed)
                    if not dry_run:
                        conn.execute(
                            f"UPDATE {table} SET {column} = ? WHERE id = ?",
                            (new, row["id"]))
                elif left:
                    result.left_alone.append(left)
                else:
                    result.already_relative += 1
        if not dry_run:
            conn.commit()
    finally:
        conn.close()
    return result


# --------------------------------------------------------------------------
# cli
# --------------------------------------------------------------------------

def report(result: Result, dry_run: bool) -> None:
    verb = "would rewrite" if dry_run else "rewrote"
    print(f"\n{result.label}")
    for line in result.changed:
        print(f"  {line}")
    for line in result.left_alone:
        print(f"  {line}")
    if not result.changed and not result.left_alone:
        print("  nothing to do")
    print(f"  {verb} {len(result.changed)}, "
          f"{result.already_relative} already relative, "
          f"{len(result.left_alone)} left alone")


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--dry-run", action="store_true",
                    help="report what would change, write nothing")
    ap.add_argument("--year", type=int,
                    help="one tracker workbook only; leaves the database alone")
    args = ap.parse_args(argv)

    # Company names carry umlauts, and a cp1252 or cp437 console cannot encode
    # all of them. Without this the report raises part-way through — after the
    # workbook has already been saved, which is the worst moment to stop.
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(errors="replace")

    root = ROOT
    results = [normalize_workbook(path, root, args.dry_run)
               for path in find_workbooks(root, args.year)]

    note, refused = "", False
    if args.year is None:
        db_path = root / "automation" / "state" / "watch.db"
        if not db_path.exists():
            note = f"\n{db_path.relative_to(root)}\n  not created yet; skipped"
        elif (pids := running_watchers()):
            refused = True
            note = (
                f"\n{db_path.relative_to(root)}\n"
                f"  REFUSED: the watcher is running (pid "
                f"{', '.join(map(str, pids))}) and would write rows back the "
                f"old way.\n"
                f"  Stop it, re-run this, then start it again:\n"
                f"    python automation/watcherctl.py stop\n"
                f"    python scripts/normalize_stored_paths.py\n"
                f"    python automation/watcherctl.py start")
        else:
            results.append(normalize_database(db_path, root, args.dry_run))

    if not results and not note:
        print("Nothing to normalize.")
        return 0
    for result in results:
        report(result, args.dry_run)
    if note:
        print(note)

    total = sum(len(r.changed) for r in results)
    print(f"\n{'Would rewrite' if args.dry_run else 'Rewrote'} {total} stored "
          f"path(s) across {len(results)} store(s).")
    if args.dry_run and total:
        print("Re-run without --dry-run to apply.")
    return 1 if refused else 0


if __name__ == "__main__":
    raise SystemExit(main())
