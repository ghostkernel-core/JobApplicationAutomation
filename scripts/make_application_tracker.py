"""Build the job-application tracker workbook for a given year.

Creates <year>/Job Applications Tracker <year>.xlsx with:
  - Applications sheet (validated dropdowns, conditional formatting, autofilter)
  - Dashboard sheet (live COUNTIF summary)
  - Lists sheet (dropdown sources, hidden)

Existing rows are preserved: if the workbook already exists the script refuses to
overwrite unless --force is passed.

Usage:
    python scripts/make_application_tracker.py            # year 2026, prefill from folders
    python scripts/make_application_tracker.py --year 2027 --no-prefill
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).resolve().parent))
from workspace_paths import ROOT, to_relative  # noqa: E402

MAX_ROWS = 500  # rows pre-armed with validation/formulas

# (header, width, dropdown list name or None)
COLUMNS = [
    ("#", 5, None),
    ("Date Applied", 13, None),
    ("Company", 26, None),
    ("Position Applied", 40, None),
    ("Location (City)", 18, None),
    ("Country", 12, None),
    ("Work Model", 13, "work_model"),
    ("Source / Channel", 18, "source"),
    ("Job Posting URL", 32, None),
    ("Language", 11, "language"),
    ("Contact Person", 20, None),
    ("Contact Email", 26, None),
    ("Status", 18, "status"),
    ("Status Updated", 15, None),
    ("Days Since Applied", 12, None),
    ("Next Action", 30, None),
    ("Follow-up Date", 14, None),
    ("Salary / Range", 16, None),
    ("Application Folder", 46, None),
    ("Notes", 50, None),
]

LISTS = {
    "status": [
        "Applied",
        "In Review",
        "Take-Home Task",
        "1st Interview",
        "2nd Interview",
        "3rd Interview",
        "Position Offered",
        "Accepted",
        "Rejected",
        "Offer Declined",
        "Withdrawn",
        "No Response",
    ],
    "work_model": ["On-site", "Hybrid", "Remote"],
    "language": ["EN", "EN + DE", "DE"],
    "source": [
        "Company Website",
        "LinkedIn",
        "Xing",
        "StepStone",
        "Indeed",
        "Glassdoor",
        "Recruiter",
        "Referral",
        "Job Board (Other)",
    ],
}

# status -> (font colour, fill colour)
STATUS_COLOURS = {
    "Applied": ("1F3864", "DDEBF7"),
    "In Review": ("1F3864", "DDEBF7"),
    "Take-Home Task": ("7F6000", "FFF2CC"),
    "1st Interview": ("7F6000", "FFE699"),
    "2nd Interview": ("7F6000", "FFD966"),
    "3rd Interview": ("7F6000", "FFC000"),
    "Position Offered": ("006100", "C6EFCE"),
    "Accepted": ("FFFFFF", "217346"),
    "Rejected": ("9C0006", "FFC7CE"),
    "Offer Declined": ("833C00", "FBE4D5"),
    "Withdrawn": ("3B3838", "D9D9D9"),
    "No Response": ("3B3838", "D9D9D9"),
}

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FOLDER_RE = re.compile(r"^(\d{4}-\d{2}-\d{2}) - (.+)$")


def discover_applications(year_dir: Path) -> list[dict]:
    """Scan <year>/<Company>/<YYYY-MM-DD> - <Role>/ and return sorted rows."""
    rows: list[dict] = []
    if not year_dir.is_dir():
        return rows
    for company_dir in sorted(year_dir.iterdir()):
        if not company_dir.is_dir() or company_dir.name.startswith((".", "_")):
            continue
        for app_dir in sorted(company_dir.iterdir()):
            if not app_dir.is_dir():
                continue
            m = FOLDER_RE.match(app_dir.name)
            if not m:
                continue
            rows.append(
                {
                    "date": dt.date.fromisoformat(m.group(1)),
                    "company": company_dir.name,
                    "role": m.group(2),
                    "folder": to_relative(app_dir),
                }
            )
    rows.sort(key=lambda r: (r["date"], r["company"], r["role"]))
    return rows


def build_lists_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Lists")
    for col, (name, values) in enumerate(LISTS.items(), start=1):
        letter = get_column_letter(col)
        ws.cell(row=1, column=col, value=name).font = Font(bold=True)
        for i, value in enumerate(values, start=2):
            ws.cell(row=i, column=col, value=value)
        ws.column_dimensions[letter].width = 20
        # named range so validation formulas stay readable
        ref = f"Lists!${letter}$2:${letter}${len(values) + 1}"
        wb.defined_names.add(__import__("openpyxl").workbook.defined_name.DefinedName(name, attr_text=ref))
    ws.sheet_state = "hidden"


def build_applications_sheet(wb: Workbook, prefill: list[dict]) -> None:
    ws = wb.create_sheet("Applications", 0)
    ws.sheet_properties.tabColor = "1F3864"

    for idx, (header, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 30

    header_index = {h: i + 1 for i, (h, _, _) in enumerate(COLUMNS)}
    col_date = get_column_letter(header_index["Date Applied"])
    col_status = get_column_letter(header_index["Status"])
    col_days = header_index["Days Since Applied"]
    col_num = header_index["#"]

    # prefilled rows
    for offset, rec in enumerate(prefill):
        r = 2 + offset
        ws.cell(row=r, column=col_num, value=offset + 1)
        ws.cell(row=r, column=header_index["Date Applied"], value=rec["date"])
        ws.cell(row=r, column=header_index["Company"], value=rec["company"])
        ws.cell(row=r, column=header_index["Position Applied"], value=rec["role"])
        ws.cell(row=r, column=header_index["Language"], value="EN")
        ws.cell(row=r, column=header_index["Status"], value="Applied")
        ws.cell(row=r, column=header_index["Application Folder"], value=rec["folder"])

    # formatting + formulas for the whole armed range
    for r in range(2, MAX_ROWS + 2):
        ws.cell(
            row=r,
            column=col_days,
            value=f'=IF({col_date}{r}="","",IF({col_status}{r}="","",TODAY()-{col_date}{r}))',
        )
        for c in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=c in (header_index["Notes"],))
        for name in ("Date Applied", "Status Updated", "Follow-up Date"):
            cell = ws.cell(row=r, column=header_index[name])
            cell.number_format = "YYYY-MM-DD"
            cell.alignment = Alignment(horizontal="center", vertical="top")
        for name in ("#", "Days Since Applied", "Work Model", "Language"):
            ws.cell(row=r, column=header_index[name]).alignment = Alignment(
                horizontal="center", vertical="top"
            )

    # dropdowns
    for header, _, list_name in COLUMNS:
        if not list_name:
            continue
        letter = get_column_letter(header_index[header])
        dv = DataValidation(
            type="list",
            formula1=f"={list_name}",
            allow_blank=True,
            showDropDown=False,
            errorTitle="Invalid entry",
            error=f"Pick one of the {header} options from the dropdown.",
        )
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{MAX_ROWS + 1}")

    # status colour coding
    status_range = f"{col_status}2:{col_status}{MAX_ROWS + 1}"
    for status, (font_colour, fill_colour) in STATUS_COLOURS.items():
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{status}"'],
                font=Font(color=font_colour, bold=True),
                fill=PatternFill("solid", start_color=fill_colour, end_color=fill_colour),
            ),
        )

    # overdue follow-up highlight
    col_follow = get_column_letter(header_index["Follow-up Date"])
    ws.conditional_formatting.add(
        f"{col_follow}2:{col_follow}{MAX_ROWS + 1}",
        CellIsRule(
            operator="lessThan",
            formula=["TODAY()"],
            font=Font(color="9C0006", bold=True),
            fill=PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE"),
        ),
    )

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{MAX_ROWS + 1}"
    ws.freeze_panes = "C2"


def build_dashboard(wb: Workbook, year: int) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = "217346"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12

    title = ws.cell(row=2, column=2, value=f"Job Applications {year}")
    title.font = Font(size=16, bold=True, color="1F3864")

    ws.cell(row=4, column=2, value="Total applications").font = Font(bold=True)
    ws.cell(row=4, column=3, value='=COUNTA(Applications!C2:C501)')

    ws.cell(row=5, column=2, value="Active (not closed)").font = Font(bold=True)
    ws.cell(
        row=5,
        column=3,
        value='=COUNTA(Applications!C2:C501)-SUMPRODUCT(COUNTIF(Applications!M2:M501,'
        '{"Rejected","Offer Declined","Withdrawn","No Response"}))',
    )

    ws.cell(row=6, column=2, value="Interviews reached").font = Font(bold=True)
    ws.cell(
        row=6,
        column=3,
        value='=SUMPRODUCT(COUNTIF(Applications!M2:M501,{"1st Interview","2nd Interview",'
        '"3rd Interview","Position Offered","Accepted"}))',
    )

    ws.cell(row=7, column=2, value="Applied in last 30 days").font = Font(bold=True)
    ws.cell(row=7, column=3, value='=COUNTIFS(Applications!B2:B501,">="&TODAY()-30)')

    ws.cell(row=8, column=2, value="Follow-ups overdue").font = Font(bold=True)
    ws.cell(
        row=8,
        column=3,
        value='=COUNTIFS(Applications!Q2:Q501,"<"&TODAY(),Applications!Q2:Q501,"<>")',
    )

    head_r = 10
    ws.cell(row=head_r, column=2, value="Status").font = HEADER_FONT
    ws.cell(row=head_r, column=3, value="Count").font = HEADER_FONT
    ws.cell(row=head_r, column=4, value="Share").font = HEADER_FONT
    for c in (2, 3, 4):
        ws.cell(row=head_r, column=c).fill = HEADER_FILL
        ws.cell(row=head_r, column=c).alignment = Alignment(horizontal="center")

    for i, status in enumerate(LISTS["status"], start=1):
        r = head_r + i
        ws.cell(row=r, column=2, value=status).border = BORDER
        count = ws.cell(row=r, column=3, value=f'=COUNTIF(Applications!M2:M501,B{r})')
        count.alignment = Alignment(horizontal="center")
        count.border = BORDER
        share = ws.cell(row=r, column=4, value=f'=IF($C$4=0,0,C{r}/$C$4)')
        share.number_format = "0.0%"
        share.alignment = Alignment(horizontal="center")
        share.border = BORDER
        font_colour, fill_colour = STATUS_COLOURS[status]
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=fill_colour)
        ws.cell(row=r, column=2).font = Font(color=font_colour, bold=True)

    note_r = head_r + len(LISTS["status"]) + 2
    ws.cell(
        row=note_r,
        column=2,
        value="Counts update automatically. Add rows in the Applications sheet "
        f"(validation and formulas are pre-armed through row {MAX_ROWS + 1}).",
    ).font = Font(italic=True, size=9, color="595959")

    ws.sheet_view.showGridLines = False


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--year", type=int, default=dt.date.today().year)
    ap.add_argument("--no-prefill", action="store_true")
    ap.add_argument("--force", action="store_true", help="overwrite an existing workbook")
    args = ap.parse_args()

    year_dir = ROOT / str(args.year)
    year_dir.mkdir(parents=True, exist_ok=True)
    out = year_dir / f"Job Applications Tracker {args.year}.xlsx"
    if out.exists() and not args.force:
        print(f"Refusing to overwrite existing {out} (pass --force).", file=sys.stderr)
        return 1

    prefill = [] if args.no_prefill else discover_applications(year_dir)

    wb = Workbook()
    wb.remove(wb.active)
    build_lists_sheet(wb)
    build_applications_sheet(wb, prefill)
    build_dashboard(wb, args.year)
    wb.active = 0
    wb.save(out)

    print(f"Wrote {out} ({len(prefill)} prefilled rows).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
