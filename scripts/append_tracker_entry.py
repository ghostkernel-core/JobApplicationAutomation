"""Append (or update) one application row in the yearly tracker workbook.

Called by the orchestrator as the last pipeline step, after final QA passes.
Idempotent: a row with the same company + position + date applied is updated in
place rather than duplicated, so re-running a pipeline never doubles an entry.

If the workbook does not exist yet it is created via make_application_tracker.py
(with --no-prefill) so validation, colours, and the dashboard are all present.

Columns are read from the sheet header row, so removing a column in Excel drops that
field from the write instead of breaking the run.

Usage:
    python scripts/append_tracker_entry.py \
        --company "ExampleCo" \
        --position "Machine Learning Engineer" \
        --date 2026-08-02 \
        --location "Frankfurt am Main" --country "Germany" \
        --folder "<workspace>/2026/ExampleCo/2026-08-02 - Machine Learning Engineer"
"""

from __future__ import annotations

import argparse
import datetime as dt
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
MAKER = ROOT / "scripts" / "make_application_tracker.py"

# header -> CLI value key. Only headers actually present in the sheet are written,
# so removing a column in Excel simply drops that field instead of breaking the run.
FIELD_BY_HEADER = {
    "Date Applied": "date",
    "Company": "company",
    "Position Applied": "position",
    "Location (City)": "location",
    "Country": "country",
    "Status": "status",
    "Status Updated": "status_updated",
    "Next Action": "next_action",
    "Follow-up Date": "follow_up",
    "Salary / Range": "salary",
    "Application Folder": "folder",
    "Notes": "notes",
}


def parse_date(value: str | None) -> dt.date | None:
    if not value:
        return None
    return dt.date.fromisoformat(value)


def tracker_path(year: int) -> Path:
    return ROOT / str(year) / f"Job Applications Tracker {year}.xlsx"


def ensure_workbook(path: Path, year: int) -> None:
    if path.exists():
        return
    subprocess.run(
        [sys.executable, str(MAKER), "--year", str(year), "--no-prefill"],
        check=True,
        cwd=str(ROOT),
    )


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--company", required=True)
    ap.add_argument("--position", required=True)
    ap.add_argument("--date", help="date applied, YYYY-MM-DD (default: today)")
    ap.add_argument(
        "--location",
        required=True,
        help='posting city, or "Remote" for a fully remote role — read it off the posting, never guess',
    )
    ap.add_argument("--country", required=True, help="posting country")
    ap.add_argument("--status", default="Applied")
    ap.add_argument("--next-action", default="")
    ap.add_argument("--follow-up", help="follow-up date, YYYY-MM-DD")
    ap.add_argument("--salary", default="")
    ap.add_argument("--folder", default="")
    ap.add_argument("--notes", default="")
    args = ap.parse_args()

    applied = parse_date(args.date) or dt.date.today()
    values = {
        "date": applied,
        "company": args.company,
        "position": args.position,
        "location": args.location,
        "country": args.country,
        "status": args.status,
        "status_updated": applied,
        "next_action": args.next_action,
        "follow_up": parse_date(args.follow_up),
        "salary": args.salary,
        "folder": args.folder,
        "notes": args.notes,
    }

    path = tracker_path(applied.year)
    ensure_workbook(path, applied.year)

    wb = load_workbook(path)
    ws = wb["Applications"]
    col = {c.value: c.column for c in ws[1] if c.value}

    c_company, c_position, c_date = col["Company"], col["Position Applied"], col["Date Applied"]

    target = None
    last_used = 1
    for r in range(2, ws.max_row + 1):
        company = ws.cell(row=r, column=c_company).value
        if not company:
            continue
        last_used = r
        existing_date = ws.cell(row=r, column=c_date).value
        if isinstance(existing_date, dt.datetime):
            existing_date = existing_date.date()
        if (
            str(company).strip().casefold() == args.company.strip().casefold()
            and str(ws.cell(row=r, column=c_position).value or "").strip().casefold()
            == args.position.strip().casefold()
            and existing_date == applied
        ):
            target = r
            break

    updated = target is not None
    row = target if updated else last_used + 1

    for header, key in FIELD_BY_HEADER.items():
        if header not in col:
            continue  # column removed from the sheet — drop the field, do not fail
        value = values[key]
        if value in ("", None):
            continue  # never blank out something already filled in by hand
        cell = ws.cell(row=row, column=col[header], value=value)
        if isinstance(value, dt.date):
            cell.number_format = "YYYY-MM-DD"
    if "#" in col:
        ws.cell(row=row, column=col["#"], value=row - 1)

    wb.save(path)
    print(
        f"{'Updated' if updated else 'Added'} row {row}: "
        f"{args.company} — {args.position} ({applied.isoformat()}) in {path}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
