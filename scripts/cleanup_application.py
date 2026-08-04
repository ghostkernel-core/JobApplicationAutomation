"""Undo one application run: delete its folder, tracker row, and scratch files.

A run that dies partway leaves debris in three places — the dated deliverable
folder, a tracker row if step 10 already fired, and payload/rasterisation
scratch under `_tmp/`. Left alone, that debris is worse than nothing: a folder
with no PDFs reads as "already applied" to a human scanning the tree, and a
tracker row for an application that was never sent is simply false.

So a failed run ends here. The watcher calls this automatically (see
`automation/watcher/builder.py`); the orchestrator calls it when an interactive
run aborts (see `CLAUDE.md`).

Deliberately narrow. It removes exactly one `<YYYY>/<Company>/<YYYY-MM-DD> -
<Role>/` folder and refuses any path that is not shaped like that or does not
sit inside the workspace. The tracker workbook is never deleted — only the one
row whose company + position + date match, using the same key
`append_tracker_entry.py` writes with.

Idempotent, and quiet about it: cleaning a run that left nothing behind is a
success with an empty report, so a caller can fire it unconditionally.

Usage:
    python scripts/cleanup_application.py \
        --folder "<workspace>/2026/ExampleCo/2026-08-02 - ML Engineer" \
        --reason "latexmk failed"

    python scripts/cleanup_application.py \
        --company ExampleCo --position "ML Engineer" --date 2026-08-02

    python scripts/cleanup_application.py --folder "<...>" --dry-run
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import shutil
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
TMP = ROOT / "_tmp"

FOLDER_RE = re.compile(r"^(\d{4}-\d{2}-\d{2})\s*-\s*(.+)$")
YEAR_RE = re.compile(r"^(19|20)\d{2}$")

# Written by make_application_tracker.py as a per-row formula. Rewriting the
# data region must not touch it, or row N ends up holding row N+1's formula.
COMPUTED_HEADERS = {"Days Since Applied"}


# --------------------------------------------------------------------------
# locating
# --------------------------------------------------------------------------

def parse_date(value: str) -> dt.date:
    return dt.date.fromisoformat(value)


def resolve_folder(company: str, position: str, applied: dt.date) -> Path | None:
    """The dated folder for this application, exact name first, then by date.

    The exact match is the normal case. The fallback exists because the folder
    name comes from the model's own reading of the posting: it may have written
    "Senior ML Engineer (m/w/d)" where the caller says "Senior ML Engineer". One
    company folder rarely holds two applications on the same day, so the date
    alone is a safe enough second key.
    """
    year_dir = ROOT / str(applied.year)
    if not year_dir.is_dir():
        return None

    wanted_company = company.strip().casefold()
    for company_dir in year_dir.iterdir():
        if not company_dir.is_dir() or company_dir.name.strip().casefold() != wanted_company:
            continue
        exact = company_dir / f"{applied.isoformat()} - {position}"
        if exact.is_dir():
            return exact
        for role_dir in sorted(company_dir.iterdir()):
            match = FOLDER_RE.match(role_dir.name) if role_dir.is_dir() else None
            if match and match.group(1) == applied.isoformat():
                return role_dir
    return None


def describe_folder(folder: Path) -> tuple[str, str, dt.date] | None:
    """(company, role, date) read back off a deliverable path."""
    match = FOLDER_RE.match(folder.name)
    if not match:
        return None
    try:
        applied = parse_date(match.group(1))
    except ValueError:
        return None
    return folder.parent.name, match.group(2), applied


def check_safe(folder: Path) -> str:
    """Empty string if this path is safe to delete, else why it is not.

    Everything here is a guard against a caller passing something it should
    not: this script runs unattended, under bypassPermissions, on a folder path
    that a model chose.
    """
    try:
        resolved = folder.resolve()
        relative = resolved.relative_to(ROOT.resolve())
    except (OSError, ValueError):
        return f"{folder} is outside the workspace"

    parts = relative.parts
    if len(parts) != 3:
        return (f"{relative} is not <YYYY>/<Company>/<date> - <Role> "
                f"({len(parts)} path segments, expected 3)")
    if not YEAR_RE.match(parts[0]):
        return f"{parts[0]} is not a year directory"
    if not FOLDER_RE.match(parts[2]):
        return f"{parts[2]} is not a dated application folder"
    if not resolved.is_dir():
        return f"{relative} is not a directory"
    return ""


# --------------------------------------------------------------------------
# the three removals
# --------------------------------------------------------------------------

def verb(dry_run: bool) -> str:
    return "would remove" if dry_run else "removed"


def remove_folder(folder: Path, dry_run: bool) -> list[str]:
    """Delete the deliverable folder, then its company folder if now empty."""
    done = [f"{verb(dry_run)} folder {folder.relative_to(ROOT)}"]
    if not dry_run:
        shutil.rmtree(folder)

    company_dir = folder.parent
    emptied = (not any(company_dir.iterdir()) if not dry_run
               else list(company_dir.iterdir()) == [folder])
    if company_dir.is_dir() and emptied:
        if not dry_run:
            company_dir.rmdir()
        done.append(f"{verb(dry_run)} empty company folder {company_dir.relative_to(ROOT)}")
    return done


def remove_tmp(company: str, applied: dt.date, stems: list[str],
               dry_run: bool) -> list[str]:
    """Payload directories for this application, and its rasterised pages.

    Payload dirs are named `<Company> <date> <Role>`, so the company+date prefix
    identifies them exactly. Page images are keyed by PDF stem instead, which is
    identity-based and therefore shared between applications — but they are a
    regenerable cache, so removing the failed run's stems costs at most a
    re-rasterisation somewhere else.
    """
    done: list[str] = []
    prefix = f"{company} {applied.isoformat()}".casefold()

    payloads = TMP / "payloads"
    if payloads.is_dir():
        for entry in sorted(payloads.iterdir()):
            if entry.is_dir() and entry.name.casefold().startswith(prefix):
                if not dry_run:
                    shutil.rmtree(entry)
                done.append(f"{verb(dry_run)} payloads {entry.relative_to(ROOT)}")

    pages = TMP / "pdf_pages"
    if pages.is_dir():
        for stem in stems:
            entry = pages / stem
            if entry.is_dir():
                if not dry_run:
                    shutil.rmtree(entry)
                done.append(f"{verb(dry_run)} page images {entry.relative_to(ROOT)}")
    return done


def remove_tracker_row(company: str, position: str, applied: dt.date,
                       dry_run: bool) -> list[str]:
    """Drop the matching row, keeping the workbook and its armed formulas intact.

    `ws.delete_rows` is wrong here: openpyxl shifts cells up without rewriting
    formulas, so the "Days Since Applied" cell that used to say `B12` lands on
    row 11 still saying `B12`. Instead the data region is read, filtered, and
    written back from row 2 — formulas, validation, conditional formatting, and
    column widths are never touched.
    """
    path = ROOT / str(applied.year) / f"Job Applications Tracker {applied.year}.xlsx"
    if not path.exists():
        return []

    try:
        from openpyxl import load_workbook
    except ImportError:
        return ["could not open the tracker: openpyxl is not installed"]

    workbook = load_workbook(path)
    if "Applications" not in workbook.sheetnames:
        return [f"no Applications sheet in {path.name}"]
    sheet = workbook["Applications"]

    header_by_col = {c.column: str(c.value) for c in sheet[1] if c.value}
    col_by_header = {h: c for c, h in header_by_col.items()}
    needed = ("Company", "Position Applied", "Date Applied")
    if not all(h in col_by_header for h in needed):
        return [f"{path.name} is missing one of {', '.join(needed)}; left untouched"]

    writable = [c for c, h in header_by_col.items() if h not in COMPUTED_HEADERS]
    c_company = col_by_header["Company"]
    c_position = col_by_header["Position Applied"]
    c_date = col_by_header["Date Applied"]

    last = 1
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(row=r, column=c_company).value:
            last = r

    def matches(r: int) -> bool:
        value = sheet.cell(row=r, column=c_date).value
        if isinstance(value, dt.datetime):
            value = value.date()
        return (
            str(sheet.cell(row=r, column=c_company).value or "").strip().casefold()
            == company.strip().casefold()
            and str(sheet.cell(row=r, column=c_position).value or "").strip().casefold()
            == position.strip().casefold()
            and value == applied
        )

    keep: list[dict[int, object]] = []
    dropped = 0
    for r in range(2, last + 1):
        if not sheet.cell(row=r, column=c_company).value:
            continue
        if matches(r):
            dropped += 1
            continue
        keep.append({c: sheet.cell(row=r, column=c).value for c in writable})

    if not dropped:
        return []
    if dry_run:
        return [f"{verb(dry_run)} {dropped} tracker row(s) from {path.name}"]

    # Assign through .value, never `cell(..., value=x)`: openpyxl reads a None
    # there as "no value supplied" and leaves the cell alone, which silently
    # turns every blanking write into a no-op.
    for offset, values in enumerate(keep):
        r = 2 + offset
        for c in writable:
            sheet.cell(row=r, column=c).value = values[c]
    for r in range(2 + len(keep), last + 1):
        for c in writable:
            sheet.cell(row=r, column=c).value = None
    if "#" in col_by_header:
        c_num = col_by_header["#"]
        for offset in range(len(keep)):
            sheet.cell(row=2 + offset, column=c_num).value = offset + 1

    workbook.save(path)
    return [f"removed {dropped} tracker row(s) from {path.name}"]


def remove_empty_year(applied: dt.date, dry_run: bool) -> list[str]:
    """Only if the year directory is now completely empty — workbook included.

    The workbook is never deleted to make this true; it is the years-long record
    and outlives any single run.
    """
    year_dir = ROOT / str(applied.year)
    if not year_dir.is_dir() or any(year_dir.iterdir()):
        return []
    if not dry_run:
        year_dir.rmdir()
    return [f"removed empty year folder {year_dir.name}"]


# --------------------------------------------------------------------------
# cli
# --------------------------------------------------------------------------

def cleanup(company: str, position: str, applied: dt.date,
            folder: Path | None, dry_run: bool = False,
            keep_tmp: bool = False, keep_tracker: bool = False) -> list[str]:
    """Everything one failed run left behind. Returns what was removed."""
    done: list[str] = []

    if folder is not None:
        # Read before deleting: the rasterisation cache is keyed by the stems of
        # the documents that are about to disappear.
        stems = sorted({f.stem for f in folder.iterdir()
                        if f.is_file() and f.suffix.lower() == ".pdf"})
        done += remove_folder(folder, dry_run)
    else:
        stems = []

    if not keep_tmp:
        done += remove_tmp(company, applied, stems, dry_run)
    if not keep_tracker:
        done += remove_tracker_row(company, position, applied, dry_run)
    if not dry_run:
        done += remove_empty_year(applied, dry_run)
    return done


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--folder", help="the dated deliverable folder to remove")
    ap.add_argument("--company", help="company as it appears in the tracker/folder")
    ap.add_argument("--position", help="role as it appears in the tracker/folder")
    ap.add_argument("--date", help="date applied, YYYY-MM-DD")
    ap.add_argument("--reason", default="", help="why the run failed, for the log line")
    ap.add_argument("--dry-run", action="store_true", help="report, remove nothing")
    ap.add_argument("--keep-tmp", action="store_true", help="leave _tmp scratch in place")
    ap.add_argument("--keep-tracker", action="store_true", help="leave the tracker row in place")
    args = ap.parse_args(argv)

    folder: Path | None = None
    if args.folder:
        folder = Path(args.folder)
        problem = check_safe(folder)
        if problem:
            print(f"Refusing to clean up: {problem}", file=sys.stderr)
            return 1
        folder = folder.resolve()
        described = describe_folder(folder)
        if described is None:
            print(f"Refusing to clean up: cannot read a date and role off {folder.name}",
                  file=sys.stderr)
            return 1
        company, position, applied = described
        # Explicit flags win over the path, which may be the fuzzy match.
        company = args.company or company
        position = args.position or position
        if args.date:
            applied = parse_date(args.date)
    else:
        if not (args.company and args.position and args.date):
            print("Give --folder, or all of --company --position --date.", file=sys.stderr)
            return 2
        company, position = args.company, args.position
        try:
            applied = parse_date(args.date)
        except ValueError:
            print(f"--date must be YYYY-MM-DD, got {args.date!r}", file=sys.stderr)
            return 2
        folder = resolve_folder(company, position, applied)
        if folder is not None:
            problem = check_safe(folder)
            if problem:
                print(f"Refusing to clean up: {problem}", file=sys.stderr)
                return 1

    done = cleanup(company, position, applied, folder,
                   dry_run=args.dry_run, keep_tmp=args.keep_tmp,
                   keep_tracker=args.keep_tracker)

    label = f"{company} - {position} ({applied.isoformat()})"
    reason = f" ({args.reason})" if args.reason else ""
    if not done:
        print(f"Nothing to clean up for {label}{reason}.")
        return 0
    print(f"Cleaned up {label}{reason}:")
    for line in done:
        print(f"  {line}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
