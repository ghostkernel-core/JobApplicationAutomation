"""One-off backfill: fill Location / Country / Status Updated on the 2026 tracker.

Every location below was read out of that application's own folder — the recipient
address block of the rendered cover letter, or the archived posting HTML — except
the four marked WEB, which were confirmed against the live posting because the
local capture carried no location text.

Matching is by (company, position, date applied), the same key append_tracker_entry.py
uses, so this is safe to re-run and will not touch a row you have edited by hand
(existing non-empty values are left alone).
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
TRACKER = ROOT / "2026" / "Job Applications Tracker 2026.xlsx"

# (company, position, YYYY-MM-DD) -> (city, country, source)
LOCATIONS = {
    ("Deutsche Börse Group", "AI Project Manager", "2026-06-16"): ("Prague", "Czechia", "posting"),
    ("Deutsche Börse Group", "Senior AI Engineer", "2026-06-16"): ("Prague", "Czechia", "posting"),
    ("Deutsche Börse Group", "Strategy AI Analyst", "2026-06-16"): ("Prague", "Czechia", "posting"),
    ("Deutsche Börse Group", "Quantitative Analyst Developer", "2026-06-17"): ("Eschborn", "Germany", "posting"),
    ("Remerge", "Machine Learning Engineer", "2026-06-17"): ("Berlin", "Germany", "posting"),
    ("Zillow Group", "Machine Learning Engineer", "2026-06-17"): ("Remote", "Germany", "posting"),
    ("Twilio", "Machine Learning Engineer", "2026-06-19"): ("Remote", "Ireland", "posting"),
    ("Deutsche Boerse Group", "IT Graduate Trainee Program - Trading Technology (fmd)", "2026-06-20"): ("Frankfurt am Main", "Germany", "posting"),
    ("Deutsche Börse AG", "IT Graduate Trainee Program - Trading Technology (fmd)", "2026-06-20"): ("Frankfurt am Main", "Germany", "posting"),
    ("Roche", "Data Scientist Machine Learning Engineer", "2026-06-24"): ("Mannheim", "Germany", "letter"),
    ("TOMRA", "Machine Learning Engineer", "2026-06-30"): ("Mülheim-Kärlich", "Germany", "WEB"),
    ("Deutsche Börse Group", "System Engineer - IT Data Operator (f-m-d)", "2026-07-01"): ("Prague", "Czechia", "letter"),
    ("Batch Robotics", "Senior Machine Learning Engineer, Robot Learning", "2026-07-07"): ("Munich", "Germany", "letter"),
    ("HiringCafe", "Founding ML AI Search Engineer", "2026-07-08"): ("Cupertino, CA", "USA", "letter"),
    ("Primal State Performance GmbH", "Applied AI & Automation Engineer (all genders)", "2026-07-12"): ("Berlin", "Germany", "letter"),
    ("Clera", "Forward Deployed Engineer", "2026-07-13"): ("Berlin", "Germany", "letter"),
    ("Synergeticon", "AI Engineer", "2026-07-13"): ("Hamburg (Finkenwerder)", "Germany", "posting"),
    ("Synergeticon", "AI-ML Developer (RAG and AI Agents)", "2026-07-13"): ("Hamburg", "Germany", "letter"),
    ("Hapag-Lloyd", "ML Engineer", "2026-07-15"): ("Gdansk", "Poland", "letter"),
    ("Deutsche Boerse Group", "Senior AI Engineer", "2026-07-21"): ("Frankfurt am Main", "Germany", "letter"),
    ("Deutsche Börse Group", "Solutions Architect - AI & Data Integration", "2026-07-21"): ("Luxembourg", "Luxembourg", "letter"),
    ("Deutsche Börse Group", "Technical AI Product Owner - Agent Chain & Agent Hub", "2026-07-21"): ("Frankfurt am Main", "Germany", "letter"),
    ("Rheinmetall", "Software Developer Payload Management", "2026-07-21"): ("Penzberg", "Germany", "letter"),
    ("Bayer", "Machine Learning Engineer", "2026-07-24"): ("Frankfurt am Main", "Germany", "letter"),
    ("Creditreform", "Machine Learning Engineer", "2026-07-24"): ("Neuss", "Germany", "letter"),
    ("Deluxe", "AI Engineer", "2026-07-24"): ("Aachen", "Germany", "letter"),
    ("Factorial", "Staff AI Engineer", "2026-07-24"): ("A Coruña", "Spain", "letter"),
    ("GlobalFoundries", "Machine Learning Engineer for Digital Manufacturing", "2026-07-24"): ("Dresden", "Germany", "letter"),
    ("Hypatos", "Technical Support Engineer", "2026-07-24"): ("Remote", "Germany", "posting"),
    ("NN Group", "AI Engineer", "2026-07-24"): ("Prague", "Czechia", "letter"),
    ("Nadara", "AI Engineer", "2026-07-24"): ("Lisbon", "Portugal", "letter"),
    ("NextexAI", "Founding Software Engineer - Backend, Cloud and AI Infrastructure", "2026-07-24"): ("Berlin", "Germany", "letter"),
    ("YPOG", "AI-ML Engineer - Generative AI + Legal AI", "2026-07-24"): ("Berlin / Hamburg / Köln / München", "Germany", "posting"),
    ("dexter health", "AI Engineer (LLM)", "2026-07-24"): ("Cologne", "Germany", "letter"),
    # CPGvision + PSignite deliberately absent: both captures are JavaScript-only
    # BambooHR shells with no location text, and public sources disagree
    # (Kraków / Naperville IL / Naples FL). Left blank rather than guessed.
}


def main() -> int:
    wb = load_workbook(TRACKER)
    ws = wb["Applications"]
    col = {c.value: c.column for c in ws[1] if c.value}

    filled, skipped, unmatched = 0, 0, []
    seen = set()

    for r in range(2, ws.max_row + 1):
        company = ws.cell(row=r, column=col["Company"]).value
        if not company:
            continue
        position = ws.cell(row=r, column=col["Position Applied"]).value
        applied = ws.cell(row=r, column=col["Date Applied"]).value
        if isinstance(applied, dt.datetime):
            applied = applied.date()

        # status date: mirror the application date where the user hasn't set one
        if not ws.cell(row=r, column=col["Status Updated"]).value and applied:
            cell = ws.cell(row=r, column=col["Status Updated"], value=applied)
            cell.number_format = "YYYY-MM-DD"

        key = (str(company), str(position), applied.isoformat() if applied else "")
        entry = LOCATIONS.get(key)
        if not entry:
            unmatched.append(f"row {r}: {company} — {position} ({key[2]})")
            continue
        seen.add(key)
        city, country, _ = entry
        if ws.cell(row=r, column=col["Location (City)"]).value:
            skipped += 1
            continue
        ws.cell(row=r, column=col["Location (City)"], value=city)
        ws.cell(row=r, column=col["Country"], value=country)
        filled += 1

    wb.save(TRACKER)

    print(f"filled {filled} rows, left {skipped} already-populated rows untouched")
    for u in unmatched:
        print("  no location on file:", u)
    for key in LOCATIONS:
        if key not in seen:
            print("  mapping unused (row missing?):", key)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
